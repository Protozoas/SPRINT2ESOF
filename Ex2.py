import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')   #Abre o arquivo
sheet = wb.get_sheet_by_name('Sheet')       #Abre a pagina pelo nome

PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}    #Dicionario com as coisas pra mudar

for rowNum in range(2, sheet.max_row()): #Laço que vai da segunda linha até a ultima
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
#O pega nome por nome e compara com as keys do dicionario, caso de True ele muda o preço

wb.save('updatedProduceSales.xlsx')        #Salva em um novo arquivo



    
