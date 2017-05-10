import openpyxl
from openpyxl.styles import Font

def preencher(sheet, num):
    fonte = Font(bold=True)  #Define um estilo para a fonte
    for i in range(2, (num+2)):
        sheet.cell(row=1, column=i).value = int(i)-1
        sheet.cell(row=i, column=1).value = int(i)-1
        sheet.cell(row=1, column=i).font = fonte   #Define o estilo da fonte na célula
        sheet.cell(row=i, column=1).font = fonte

def multiplicar(sheet, num):
    for i in range(2,(num+2)):
        for j in range(2,(num+2)):
            sheet.cell(row=i, column=j).value = (i-1)*(j-1)

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
print('Digite um numero:')
N = int(input())
preencher(sheet,N)
multiplicar(sheet,N)
wb.save('teste.xlsx')
